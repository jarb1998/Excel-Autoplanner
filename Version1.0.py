import openpyxl
from openpyxl.styles import *
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import column_index_from_string
import xlsxwriter
import xlsxwriter.utility


file1 = openpyxl.load_workbook('{blank} Timetable Planner TEST V1.xlsx')
sheetNam = file1.sheetnames
curSheet = file1[sheetNam[0]]

with open('Overwrite log.txt', 'w') as error_log:
    error_log.write('')


def range_group_check(*group):
    """ Takes n number arguments of groups and ranges through to the maximum number of n groups two values are
    appended to lists if they are not merged and have an internal name value found by the get_group_name(): function.
     (x being the first value of row header and x-2 being last value of row header)
     The list is zipped and the first num of list two is nonsensical as this row num is before the first row header and
     so is deleted from list. To get the final row num max_row is used and appended to list the zipped list of these are
     then used as a parameter through dat_cal():"""
    listone = []
    listtwo = []
    max_num = len(list(group))
    for s in list(group[0:max_num]):
        for x in range(1, curSheet.max_row):
            if type(curSheet['B%d' % (x)]).__name__ != "MergedCell" and (curSheet['B%d' % (x)]).internal_value == s:
                listone.append(x)
                listtwo.append(x - 2)

    listtwo.pop(0)
    listtwo.append(curSheet.max_row)
    var = list(zip(listone, listtwo))
    return var

def get_group_name():
    """
    Takes the range of the Excel Sheet and determines each merged cell by it's top Left Value (Name)
    the cell value is appended to the list group_names and any with No value are filtered out.
    This is then put through the range_group_check function.
    """
    group_names = []

    for n in range(1, curSheet.max_row):
        if type(curSheet['B%d' % (n)]).__name__ != "MergedCell":
            a = curSheet['B%d' % (n)]
            b = a.internal_value
            group_names.append(b)
            group_names = [i for i in group_names if i is not None]
    return range_group_check(*group_names)

def cell_evaluater(cell_obj):
    """
    The cell evaluator function iterates through all items in pri_col and if they are greater than 500
    it saves that value to col_lis dictionary, wherein the value of the cell is the key.
    It also notes the colour of the cell and saves the contents of the cell and its hex colour to col_lis
    list under the appropriate dictionary key.
    If it finds col_lis cell containing col_lis value > 500 it will look across the same row under columns specified by
    *sec_col, and add the contents of this cell and its color to the appropriate key in the dictionary.
    """
    org_cell_con = []
    cell_val = cell_obj.value
    if cell_val is not None:
        cell_val = cell_obj.value
        cell_color = cell_obj.fill.start_color.index
        cell_coords = cell_obj.coordinate
        org_cell_con = org_cell_con + [cell_val, cell_color, cell_coords]
    return org_cell_con


def merge_ranges(class_dict):
    """
    This function filters through the list of merged cells (ranges_index) and adds only merge ranges which appear in the
    class_dict values. It then passes said list to the filter match function, which adds the new coordinates into the
    lists
    """
    ranges_index = merge_ids()
    dict_vals = list(class_dict.values())
    range_mod = [coord[0:coord.index(':')] for coord in ranges_index]
    merge_match = []
    for i in dict_vals:
        if i[2] in range_mod:
            index = range_mod.index(i[2])
            merge_span = ranges_index[index]
            merge_match.append([i[2:len(i)], merge_span])
    return merge_match


def merge_ids():
    """
    This returns a alphabetically sorted a list of all merged cells in the current excel sheet

    """
    mer_ranges = curSheet.merged_cells.ranges
    merge_where = [str(val) for val in mer_ranges]
    return sorted(merge_where)

def staff_index(var):
    """
    This iterates over the list of staff(which is gotten from the var parameter) and then finds the row number
    associated with each member of staff, returning a dictionary with this information
    """
    staff_row = {}
    for i in curSheet['C%d:C%d' % (var[0][0], var[0][1])]:
        i_len = len(i)
        i_val = i[0].value
        if type(i_val) != type(None):  # this filters out none type results
            i_val_list = i_val.split()
            sep_initials = i_val_list[i_len + 1]
            staff_row[sep_initials] = i[0].row  # this sets the key to the initals
    return staff_row

def cell_writer(column_letter, row_num, class_dict, cell_class_val):
    """
    This writes all information to the destination cell (which is dictated by the parameters column_letter and row_num),
    it retrieves the text information from class_dict and what colour to fill from class_dict, the desitnation cell is
    then append to the appropriate entry in the dictionary

    The if statement checks whether something has already been written to that cell, and if it has it will update
    the overwrite log to let the user know there was an overwrite and that coordinate.
    """

    cur_val = curSheet[column_letter + str(row_num)].value
    coords = column_letter + str(row_num)
    if type(cur_val) != type(None):  # this writes in to the error log overwrites
        with open('Overwrite log.txt', 'a') as error_log:
            cur_val = cur_val.replace('\n', ' ')
            new_val = class_dict[cell_class_val][0].replace('\n', ' ')
            str1 = 'class %s at %s was overwritten with %s' % (cur_val, coords, new_val) + '\n'
            error_log.write(str1 + '\n')
        curSheet[coords] = class_dict[cell_class_val][0]
        curSheet[coords].fill = PatternFill(start_color=class_dict[cell_class_val][1],
                                            fill_type='solid')
        class_dict[cell_class_val].append(column_letter + str(row_num))
        return class_dict
    else:
        curSheet[coords] = class_dict[cell_class_val][0]
        curSheet[coords].fill = PatternFill(start_color=class_dict[cell_class_val][1],
                                            fill_type='solid')
        class_dict[cell_class_val].append(coords)
        return class_dict

def dat_cal():
    """
    This is the main function of the program, it iterates through all the classes and will then write said information
    to the relevant new cell based on who is supposed to staff that class.
    It relies on and calls on the functions cell_writer, staff_index, merge_ranges, filter_match, merge_ids,
    cell_evaluator, get_group_name, and range_group_check to process the information so it knows what to write to
    were
    """

    var = get_group_name()
    staff_dict = staff_index(var)
    class_dict = {}
    count = 0
    num_max_col = curSheet.max_column
    col_name_end = xlsxwriter.utility.xl_col_to_name(num_max_col)
    for j in range(var[1][0], var[1][1], 2):
        cell_investigate = curSheet['D%d:%s%d' % (j, col_name_end, j)]
        for i in cell_investigate[0]:
            cell_class_val = i.value
            if type(cell_class_val) != type(None):  # This filters out blank cells
                class_dict[cell_class_val] = cell_evaluater(i)
                class_list = cell_class_val.split()

                teach_initals = class_list[len(class_list) - 1]
                if teach_initals in staff_dict.keys():  # this filters out blocks like self study
                    new_row_num = staff_dict[teach_initals]
                    column_letter = get_column_letter(i.column)
                    class_dict = cell_writer(column_letter, new_row_num, class_dict, cell_class_val)
            row_num = i.row + 1
            col_num = i.column
            col_let = get_column_letter(col_num)
            under_cell_val = curSheet[col_let + str(row_num)].value

            if type(under_cell_val) != type(None):
                if under_cell_val not in class_dict.keys():
                    class_dict[under_cell_val] = cell_evaluater(curSheet[col_let + str(
                        row_num)])
                else:
                    under_cell_val = under_cell_val + str(count)
                    class_dict[under_cell_val] = cell_evaluater(curSheet[col_let + str(row_num)])

                class_list = under_cell_val.split()
                teach_initals = class_list[len(class_list) - 1]
                if teach_initals in staff_dict.keys():
                    new_row_num = staff_dict[teach_initals] + 1
                    class_dict = cell_writer(col_let, new_row_num, class_dict, under_cell_val)
    merge_coords = merge_ranges(class_dict)
    return merge_coords, class_dict


merge_coords, class_dict = dat_cal()

file1.save('{blank} Timetable Planner TEST(ouput).xlsx')

### Cell Merging ###
file1 = openpyxl.load_workbook('{blank} Timetable Planner TEST(ouput).xlsx')
sheetNam = file1.sheetnames
curSheet = file1[sheetNam[0]]


def merge_offset(merge_coords, multCoords=True):
    """
    This function gets the coordinates that need to be merged from merge_coord and finds out over what area they
    were merged, e.g. if they were a 1x2 or 2x2 merged cell, assuming its parameter multCoords= True.

    It then returns this as list of how much to merge across the way and down the way. This is plus the current cell its
    on, so when it returns [1,0] it actually means merge two cells across the way and merge 0 cells down the way.
    """

    str1 = merge_coords
    row_nums = ''.join(digi for digi in str1 if digi.isdigit() or digi == ':')
    col_lets = ''.join(char for char in str1 if char.isalpha() or char == ':')
    col_lis = col_lets.split(':')
    row_lis = row_nums.split(':')
    if multCoords == True:
        col_offset = column_index_from_string(col_lis[1]) - column_index_from_string(col_lis[0])
        row_offset = int(row_lis[1]) - int(row_lis[0])
    else:
        return col_lis + row_lis
    return [col_offset, row_offset]


for j in merge_coords:
    if len(j[0])> 1:
        merge_where = merge_offset(j[0][1], multCoords=False)
    else:
        continue
    offset = merge_offset(j[1])
    start_col_num = column_index_from_string(merge_where[0])
    curSheet.merge_cells(start_row=int(merge_where[1]), start_column=start_col_num,
                         end_row=int(merge_where[1]) + int(offset[1]), end_column=start_col_num + offset[0])


file1.save('{blank} Timetable Planner TEST(ouput).xlsx')


