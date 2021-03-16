import openpyxl
from openpyxl.styles import *
file1 = openpyxl.load_workbook('Max-reformatted - Colour corrected.xlsx')
sheetNam = file1.sheetnames
curSheet = file1[sheetNam[0]]

'''
The cell evaluator function iterates through all items in pri_col and if they are greater than 500
it saves that value to a dictionary, wherein the value of the cell is the key.
It also notes the colour of the cell and saves the contents of the cell and its hex colour to a 
list under the appropriate dictionary key.
If it finds a cell containing a value > 500 it will look across the same row under columns specified by
*sec_col, and add the contents of this cell and its color to the appropriate key in the dictionary.
'''
def cell_evaluater(pri_col, *sec_col):
    col_lis = [pri_col, *sec_col] #This specifies all the cols to look at
    cell_dict = {}
    for i in curSheet[col_lis[0]][1:(curSheet.max_row + 1)]: #This iterates through all the cells in pri-col which have data
        try:
            if i.value.isdigit():
                cell_con = float(i.value)
        except:
            cell_con = i.value
        cell_row = i.row
        if cell_con > 500:
            cell_dict[cell_con] = []
            for j in col_lis:
                cell_val = curSheet[j+str(cell_row)].value #this gets the value in the cell
                cell_color = curSheet[j+str(cell_row)].fill.start_color.index #this gets the hex code for the cell colour
                print(cell_color)
                cell_dict[cell_con].append([cell_val, cell_color]) #this appends a list containing cell values and colour to a list containg the information for all relevant columns
    return cell_dict


'''
This function takes a user input to specify which primary column to evaluate and then the columns to take the value
from if the cell in the primary column returns true.
'''
def dat_cal():
    a = 'G'#input('Select primary column: ')
    b = 'B'#input('Second associated column: ')
    c = 'A'#input('Third associated column: ')
    cell_dict = cell_evaluater(a, b, c)
    dict_keys = cell_dict.keys() #this will evalute all the relevant cells
    row_num = 2
    print(cell_dict)
    for i in dict_keys:
        for j in range(len(cell_dict[i])): #this goes through every list in the list assigned to the relevant key
            print(cell_dict[i][j][1])
            if j == 0: #This relates to primary column information to carry across
                curSheet['K'+str(row_num)].fill = PatternFill(start_color=cell_dict[i][j][1], fill_type='solid')
                curSheet['K'+str(row_num)] = cell_dict[i][j][0]

            elif j == 1: #This and j==2 writes the relevant information from assciated seconday columns.
                curSheet['J' + str(row_num)].fill = PatternFill(start_color=cell_dict[i][j][1], fill_type='solid')
                curSheet['J' + str(row_num)] = cell_dict[i][j][0]

            elif j == 2:
                curSheet['I' + str(row_num)].fill = PatternFill(start_color=cell_dict[i][j][1], fill_type='solid')
                curSheet['I' + str(row_num)] = cell_dict[i][j][0]
        row_num += 1 #this shifts to the next row in the primary column
dat_cal()

# Temporary example Cell merging
ws.merge_cells('A2:D3')

# Cell alignment
curCell = ws['A2']  # aligns the text in the merged cells by calling the coordinates of the top left cell.
curCell.alignment = Alignment(horizontal='center', vertical='center')

file1.save('Highlighted information.xlsx')
